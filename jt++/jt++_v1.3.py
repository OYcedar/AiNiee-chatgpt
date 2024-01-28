import re
import os
import json
import pandas as pd
import traceback
from ruamel.yaml import YAML
import openpyxl

# t++标红标蓝行需要找对应code
# code对应地址
# "Plugin Command": "356","Control Variables": "122","Script": "655",
redcode = ['356', '655', '122']
# "Comment": "108","Comment More": "408",
bluecode = ['108', '408']
bluedir = [r'System.json\switches', r'System.json\variables']
# "Show Choices": "102","Show Text Attributes": "101","Show Text": "401","Show Scrolling Text": "405",
# "Show Scrolling Text Attributes": "105","Change Actor Name": "320","Change Actor Nickname": "324",
# "Choice": "402"应该和102一起来
textcode = ['-1', '401', '101', '102','105','405','320',"324"]  # 需要被翻译的大概只有这些，-1是没有code的
# "Label": "118","Jump to Label": "119","Conditional Branch": "111","Show Picture": "231",
emptycode = ['357', '657', '111', '118', '119'] # t++没有提取的，应该不止

RPG_CODE={
  "Empty": "0",
  "Show Text Attributes": "101",
  "Show Choices": "102",
  "Input Number": "103",
  "Select Key Item": "104",
  "Show Scrolling Text Attributes": "105",
  "Comment": "108",
  "Conditional Branch": "111",
  "Loop": "112",
  "Break Loop": "113",
  "Exit Event Processing": "115",
  "Call Common Event": "117",
  "Label": "118",
  "Jump to Label": "119",
  "Control Switches": "121",
  "Control Variables": "122",
  "Control Self Switch": "123",
  "Control Timer": "124",
  "Change Gold": "125",
  "Change Items": "126",
  "Change Weapons": "127",
  "Change Armor": "128",
  "Change Party Member": "129",
  "Change Battle BGM": "132",
  "Change Battle End ME": "133",
  "Change Save Access": "134",
  "Change Menu Access": "135",
  "Change Encounter": "136",
  "Change Formation Access": "137",
  "Change Window Color": "138",
  "Transfer Player": "201",
  "Set Vehicle Location": "202",
  "Set Event Location": "203",
  "Scroll Map": "204",
  "Set Move Route": "205",
  "Get on/off Vehicle": "206",
  "Change Transparency": "211",
  "Show Animation": "212",
  "Shot Balloon Icon": "213",
  "Erase Event": "214",
  "Change Player Followers": "216",
  "Gather Followers": "217",
  "Fadeout Screen": "221",
  "Fadein Screen": "222",
  "Tint Screen": "223",
  "Flash Screen": "224",
  "Shake Screen": "225",
  "Wait": "230",
  "Show Picture": "231",
  "Move Picture": "232",
  "Rotate Picture": "233",
  "Tint Picture": "234",
  "Erase Picture": "235",
  "Set Weather Effects": "236",
  "Play BGM": "241",
  "Fadeout BGM": "242",
  "Save BGM": "243",
  "Replay BGM": "244",
  "Play BGS": "245",
  "Fadeout BGS": "246",
  "Play ME": "249",
  "Play SE": "250",
  "Stop SE": "251",
  "Play Movie": "261",
  "Change Map Display": "281",
  "Change Tileset": "282",
  "Change Battle Back": "283",
  "Change Parallax Back": "284",
  "Get Location Info": "285",
  "Battle Processing": "301",
  "Shop Processing": "302",
  "Name Input Processing": "303",
  "Change HP": "311",
  "Change MP": "312",
  "Change State": "313",
  "Recover All": "314",
  "Change EXP": "315",
  "Change Level": "316",
  "Change Parameters": "317",
  "Change Skills": "318",
  "Change Equipment": "319",
  "Change Actor Name": "320",
  "Change Actor Class": "321",
  "Change Actor Graphic": "322",
  "Change Vehicle Graphic": "323",
  "Change Actor Nickname": "324",
  "Change Actor Profile": "325",
  "Change Enemy HP": "331",
  "Change Enemy MP": "332",
  "Change Enemy State": "333",
  "Enemy Recover All": "334",
  "Enemy Appear": "335",
  "Enemy Transform": "336",
  "Show Battle Animation": "337",
  "Force Action": "339",
  "Abort Battle": "340",
  "Open Menu Screen": "351",
  "Open Save Screen": "352",
  "Game Over": "353",
  "Return to Title Screen": "354",
  "Script Header": "355",
  "Plugin Command": "356",
  "Show Text": "401",
  "Choice": "402",
  "Choice Cancel": "403",
  "Choices End": "404",
  "Show Scrolling Text": "405",
  "Comment More": "408",
  "Else": "411",
  "Branch End": "412",
  "Repeat Above": "413",
  "If Win": "601",
  "If Escape": "602",
  "If Lose": "603",
  "Battle Processing End": "604",
  "Shop Item": "605",
  "Script": "655"
}


class Jr_Tpp():
    def __init__(self,config:dict,path:str=False):
        self.config=config
        self.ProgramData= {} # 翻译工程数据,键为文件名，值为DataFrame，列为['原文','译文','地址','标签']，同时设置原文为索引
        self.BlackDir=config['BlackDir']   # 地址黑名单，自动打标签“BlackDir”
        self.BlackFiles=config['BlackFiles']    # 黑名单文件，不读取这些文件，需要是文件全名（也不会有人把Map加黑名单吧）
        self.BlackCode=config['BlackCode']      # 效果同blackdir，只不过这个是code
        self.NameWithout=config['NameWithout']  #   对这些字段搜索反选后，打Name标签
        self.codewithnames=config['codewithnames']  # dnb用，包裹文件名的标识符
        self.ReadCode=config['ReadCode']    # 只读取这些code的文本
        if path:
            self.load(path) # 从工程文件加载

####################################读取和注入游戏文本，保存与加载翻译工程，导入翻译结果等基本功能###################################
    # 用openpyxl读xlsx，因为用pandas会把'=xxx'的字符串读成NaN，而且解决不了
    def __Readxlsx(self, name):
        try:
            # 打开Excel文件
            workbook = openpyxl.load_workbook(name)
            # 获取所有工作表的名称
            sheet_names = workbook.sheetnames
            # 选择第一个工作表
            worksheet = workbook[sheet_names[0]]
            # 读取列名
            column_names = [cell.value for cell in worksheet[1] if cell.value is not None]
            # 检查是否只有列名
            if len(column_names) == 0:
                # 创建空DataFrame
                df = pd.DataFrame(columns=column_names)
            else:
                # 读取数据
                data = []
                for row in worksheet.iter_rows(min_row=2, values_only=True):
                    data.append(row)

                # 创建DataFrame
                df = pd.DataFrame(data, columns=column_names)
            # 关闭Excel文件
            workbook.close()
            return df
        except Exception as e:
            print(traceback.format_exc())
            print(e)
            print('请关闭所有xlsx文件再试')
    # 读取json文件中含中日字符的字符串，并记录其地址。
    # 输入json文件的内容，返回其中所有文本组成的list
    def __ReadFile(self,data,FileName:str,code:int =False) -> list:
        res=[]
        tp = type(data)
        if tp == dict:
            for key in data.keys():
                # FileName用来记录地址
                code = data.get('code', False)
                res+=self.__ReadFile(data[key],FileName+'\\'+key,code)

        elif tp == list:
            for i in range(0,len(data)):
                res+=self.__ReadFile(data[i],FileName+'\\'+str(i),code)
        # 是字符串，而且含中日字符(System.json\gameTitle不论是否含中日字符，都进
        elif tp==str and (re.search(r'[\u4e00-\u9fa5\u3040-\u309f\u30a0-\u30ff\u4e00-\u9fa5]',data)
                          or r'System.json\gameTitle' in FileName) :
            if not code:code='-1'
            # 只有特定code才读取
            if str(code) in self.ReadCode:
                res.append([data,'',FileName,'',str(code)])
        return res
    # 读取文件夹路径，返回包括其子文件夹内的所有文件名
    def __ReadFolder(self,dir:str) -> list:
        res=[]
        if os.path.isdir(dir):
            FileList = os.listdir(dir)
            for name in FileList:
                temp=dir+'\\'+name
                if os.path.isfile(temp):
                    res+=[temp]
                elif os.path.isdir(temp):
                    res+=self.__ReadFolder(temp)
            return res
        else:return [dir]
    # 判断地址是否为黑名单地址
    def __IfBlackDir(self,Dir:str) ->bool:
        for blackdir in self.BlackDir:
            # 如果有任意一个黑名单遍历后仍为True，则说明地址为黑名单地址，break出来
            dirsig = True
            for i in blackdir.split('*'):
                if i not in Dir:
                    dirsig = False
                    break
            if dirsig:
                return True
        return False
    # 去除DataFrame中重复的行，将重复行的地址和code添加到被保留的行中,在地址为黑名单地址或code为黑名单code的情况下，不保留其数据
    def __RemoveDuplicated(self,data:pd.DataFrame) -> pd.DataFrame:
        if '地址' and 'code' in data.columns:
            a = data[~data.index.duplicated()].copy()  # 去除重复行的
            b = data[data.index.duplicated()].copy()  # 仅含重复行的
            a = {'a': a}
            b = {'b': b}
            for index in a['a'].index:
                if index in b['b'].index:
                    Dir = list(b['b'][b['b'].index == index]['地址'])
                    code = list(b['b'][b['b'].index == index]['code'])
                    black=False
                    if self.__IfBlackDir(Dir):black=True
                    if code in self.BlackCode: black = True
                    if not black:
                        for i in range(0,len(Dir)):
                            a['a'].loc[index,'地址']+='☆↑↓'+Dir[i]
                            a['a'].loc[index, 'code'] += ','+code[i]
            return a['a']
        else:
            return data[~data.index.duplicated()]
    # 将ReadFile得到的数据转化为DataFrame
    def __toDataFrame(self,data:list) -> pd.DataFrame:
        DataFrame = pd.DataFrame(data, columns=['原文', '译文', '地址', '标签','code'])
        DataFrame.index = list(DataFrame['原文'])
        DataFrame=self.__RemoveDuplicated(DataFrame)# 去除原文重复的行
        return DataFrame
    # 将后缀json和xlsx互相转化
    def __nameswitch(self,name):
        name = name.split('\\')[-1]
        resname = ''
        temp = name.split('.')
        for i in temp[:-1]:
            resname += i + '.'
        if 'xlsx' in name:
            resname += 'json'
        else:
            resname += 'xlsx'
        return resname
    # 按照Dir逐级读取文件内容，直到读到untrs，将其替换为trsed，然后逐级返回
    def __WriteFile(self,data,untrs:str,trsed:str,Dir:list):
        # 获取文本在文件内的地址
        if type(data)==list:
            i=int(Dir[0])
            data[i]=self.__WriteFile(data[i],untrs,trsed,Dir[1:])
        elif type(data)==dict:
            data[Dir[0]]=self.__WriteFile(data[Dir[0]],untrs,trsed,Dir[1:])
        elif type(data)==str and len(Dir)==0:
            if data==untrs:
                data=trsed
            else:
                print(f'原文\"{data}\"不匹配')
        return data
    # 检查译文中是否存在空数据，若存在，用原文填充，并输出提示
    def __CheckNAN(self):
        nanlist=[]
        for name in self.ProgramData.keys():
            DataFrame=self.ProgramData[name]
            NANFrame=DataFrame[DataFrame['译文'].isnull()]
            if len(NANFrame):
                indexlist=list(NANFrame.index)
                nanlist+=indexlist
                for index in indexlist:
                    DataFrame.loc[index,'译文']=index
            self.ProgramData[name]=DataFrame
        if len(nanlist):
            print('以下原文没有对应译文，恢复为原文')
            for i in nanlist:
                print(i)
    # 从游戏读取文本,参数为游戏目录，自动标签黑名单地址,并对其应用原文
    def ReadGame(self,GameDir:str):
        Files=self.__ReadFolder(GameDir)
        for File in Files:
            # 只读取data内的json文件
            if '\\data\\' in File and 'json' in File:
                name=File.split('\\')[-1]
                # 黑名单文件不读取
                if name not in self.BlackFiles:
                    print(f'正在读取{name}')
                    with open(File, 'r', encoding='utf8') as f:
                        data = json.load(f)
                    TextDatas=self.__ReadFile(data,name)
                    self.ProgramData.update({name:self.__toDataFrame(TextDatas)})
        print('########################读取游戏完成########################')
        ## 标签黑名单地址,并对其应用原文
        # self.LabelBlackDir()
        # 有可能有些文本有多个地址和code，但只有其中一个是黑的，所以不再事先标记黑名单，而在注入时，对每个地址/code单独判断是否是黑的
    # 注入翻译到游戏,BlackLabel为不注入的标签list，默认为'BlackDir',BlackCode默认self.BlackCode
    def InjectGame(self,GameDir:str,path:str,BlackLabel:list=False,BlackCode:list=False):
        self.__CheckNAN()
        if not BlackLabel:
            BlackLabel=['BlackDir']
        if not BlackCode:
            BlackCode=self.BlackCode
        Files = self.__ReadFolder(GameDir)
        for File in Files:
            # 只读取data内的json文件
            if '\\data\\' in File and 'json' in File:
                name = File.split('\\')[-1]
                # 黑名单文件不写入
                if name not in self.BlackFiles:
                    print(f'正在写入{name}')
                    with open(File, 'r', encoding='utf8') as f:
                        data = json.load(f)
                    # 写入翻译
                    if name in self.ProgramData.keys():
                        DataFrame=self.ProgramData[name]
                        for untrs in DataFrame.index:
                            trsed=DataFrame.loc[untrs,'译文']
                            Dirlist=DataFrame.loc[untrs,'地址'].split('☆↑↓')
                            codelist=DataFrame.loc[untrs,'code'].split(',')
                            labellist=DataFrame.loc[untrs,'标签'].split(',')
                            black=False
                            for label in labellist:
                                if label in BlackLabel:
                                    black=True
                            for i in range(0,len(Dirlist)):
                                Dir=Dirlist[i]
                                code=codelist[i]
                                # 标签，地址和code都不是黑的才写入
                                if not black and code not in BlackCode and not self.__IfBlackDir(Dir):
                                    Dir=Dir.split('\\')
                                    # 写入翻译
                                    data=self.__WriteFile(data,untrs,trsed,Dir[1:])
                    else:
                        print(f'{name}不在工程文件中，工程文件与游戏是否匹配')
                    # 获取文件输出路径
                    outputpath=(path+'\\data\\'+File.split('\\data\\')[-1]).lstrip('\\')
                    # 获取并创建从path到outputpath的路径
                    datadir=(outputpath.replace(name,'').replace(path,'').strip('\\')).split('\\')
                    temp=path.rstrip('\\')
                    for i in datadir:
                        temp+='\\'+i
                        if not os.path.exists(temp.strip('\\')): os.mkdir(temp.strip('\\'))
                    # 输出文件
                    out = json.dumps(data, ensure_ascii=False)
                    with open(outputpath, 'w', encoding='utf8') as f1:
                        print(out, file=f1)
        print('########################写入游戏完成########################')
    # 获取翻译工程内的文件名,Mapxxx.json合并为Mapxxx~XXX.json
    def GetFileNames(self) -> list:
        namelist=list(self.ProgramData.keys())
        mapname=[]
        for name in namelist.copy():
            if 'map' in name.lower() and 'info' not in name.lower():
                namelist.remove(name)
                mapname.append(int(name.replace('Map','').replace('.json','')))
        mapname=sorted(mapname)
        if int(mapname[0])<10:
            mapname[0]=f'00{mapname[0]}'
        elif int(mapname[0])<100:
            mapname[0]=f'0{mapname[0]}'
        if int(mapname[-1])<10:
            mapname[-1]=f'00{mapname[-1]}'
        elif int(mapname[-1])<100:
            mapname[-1]=f'0{mapname[-1]}'
        namelist.append(f'Map{mapname[0]}~{mapname[-1]}.json')
        return namelist
    # 导出单个文件
    def ToXlsx(self,name:str,path:str):
        # 文件名后缀改为xlsx
        print(f'正在保存{self.__nameswitch(name)}')
        outputname = self.__nameswitch(name)
        data=self.ProgramData[name]
        # 将标签转化为字符串
        try:
            data.to_excel(path + '\\' + outputname, index=False)
        except Exception as e:
            print(traceback.format_exc())
            print(e)
            input('保存失败，请关闭所有xlsx文件后再次尝试')
    # 保存/导出工程,数据保存为xlsx，设置保存为json
    def Save(self,path:str):
        if not os.path.exists(path+'\\data'): os.mkdir(path+'\\data')
        for name in self.ProgramData.keys():
            self.ToXlsx(name,path+'\\data')
        out = json.dumps(self.config, indent=4, ensure_ascii=False)
        with open(path+'\\'+'config.json', 'w', encoding='utf8') as f1:
            print(out, file=f1)
        print('########################保存工程完成########################')
    # 导入翻译，从json导入,路径需指定到json文件。可指定某几个文件(list格式)，namelist为False则全选
    # trsdata和path二选一
    def InputFromJson(self,trsdata:dict=False,path:str=False,namelist:list=False):
        if not trsdata:
            try:
                with open(path, 'r', encoding='utf8') as f:
                    trsdata = json.load(f)
            except Exception as e:
                print(traceback.format_exc())
                print(e)
                input(f'读取{path}失败,请确保json文件头格式正确')
        # 全选
        if not namelist:
            namelist=self.ProgramData.keys()
        # 对每一个文件都遍历一次json文件（有可能有重复原文）
        for name in namelist:
            DataFrame=self.ProgramData[name]
            for untrs in trsdata.keys():
                if untrs in DataFrame.index:
                    DataFrame.loc[untrs,'译文']=trsdata[untrs]
            self.ProgramData[name]=DataFrame
    # 从DataFrame导入翻译，指定namelist.默认有列名，第一列为原文，第二列为译文。
    # samefile为true时，不搜索原文，直接用译文列覆盖原数据
    def InputFromDataFrame(self, data:pd.DataFrame, namelist:list=False, samefile=False):
        # 格式化data
        col=list(data.columns)
        if len(col)>=2:
            col[0]='原文'
            col[1]='译文'
            data.columns=col
            data.index=list(data['原文'])
            data['译文']=data['译文'].fillna('')
            data=self.__RemoveDuplicated(data)# 去除原文重复的行
            if not namelist:
                namelist=self.ProgramData.keys()
            for name in namelist:
                DataFrame=self.ProgramData[name]
                if not samefile:
                    for untrs in data.index:
                        if untrs in DataFrame.index:
                            DataFrame.loc[untrs,'译文']=data.loc[untrs,'译文']
                else:
                    DataFrame['译文']=data['译文']
                self.ProgramData[name]=DataFrame
    # 导入翻译，从xlsx导入，可指定到文件夹也可指定到xlsx文件。
    def InputFromeXlsx(self, path:str, namelist:list=False, samefile=False):
        FileNames=self.__ReadFolder(path) # 读取path内文件路径
        for file in FileNames:
            if 'xlsx' in file:
                print('正在读取{}'.format(file.split('\\')[-1]))
                try:
                    data = self.__Readxlsx(file)
                except Exception as e:
                    print(traceback.format_exc())
                    print(e)
                    input(f'读取{path}失败')
                if not samefile:
                    self.InputFromDataFrame(data,namelist)
                else:
                    # 获取文件的json文件名
                    name=file.split('\\')[-1]
                    jsonname =self.__nameswitch(file)
                    # 检查是否存在该json文件
                    if jsonname in self.ProgramData.keys():
                        self.InputFromDataFrame(data,[jsonname],True)
                    else:
                        print(f'{name}没有对应的json文件，已跳过导入该文件')
    # 读取工程,只从xlsx读取，path内需包含一个config
    # xlsx必须含['原文','译文','地址','code']五列，读取时自动将xlsx名转化为json名，并作为键
    def load(self,path:str):
        self.ProgramData={} # 清空工程数据
        FileNames = self.__ReadFolder(path) # 读取path内文件路径
        for file in FileNames:
            print('正在加载{}'.format(file.split('\\')[-1]))
            if 'xlsx' in file:
                data = self.__Readxlsx(file)
                # 检查xlsx格式是否正确
                if not list(data.columns)==['原文','译文','地址','标签','code']:
                    print(f'{file}文件列名不为[\'原文\',\'译文\',\'地址\',\'标签\',\'code\']，读取失败')
                    continue
                # 设置索引
                data.index = list(data['原文'])
                # 填充译文和标签空数据
                data['译文'] = data['译文'].fillna('')
                data['标签'] = data['标签'].fillna('')
                # 去除原文重复的行
                data=self.__RemoveDuplicated(data)
                # 获取文件的json文件名
                jsonname =self.__nameswitch(file)
                self.ProgramData.update({jsonname:data})
        print('########################加载工程完成########################')
############################################输出，增减标签，按条件搜索及其配套操作函数###########################################
    # 添加标签,target格式为{文件名:[索引（原文)]}
    def addlabel(self,target:dict,label:str):
        for name in target.keys():
            for index in target[name]:
                if index in self.ProgramData[name].index:
                    # 如果标签已经存在，不重复添加
                    if label not in self.ProgramData[name].loc[index,'标签']:
                        self.ProgramData[name].loc[index,'标签']=(self.ProgramData[name].loc[index,'标签']+','+label).lstrip(',')
                else:
                    print(f'在{name}中没有找到{name}:\"{index}\"行')
        print(f'添加标签\"{label}\"完成')
    # 去除标签
    def removelabel(self,target:dict,label:str):
        for name in target.keys():
            for index in target[name]:
                if index in self.ProgramData[name].index:
                    # 确保标签存在
                    if label in self.ProgramData[name].loc[index,'标签']:
                        # 先在最左边加上一个逗号，然后删掉逗号+label，然后把最左边逗号去掉
                        self.ProgramData[name].loc[index,'标签']=(','+self.ProgramData[name].loc[index,'标签']).replace\
                            ((','+label),'').lstrip(',')
                    else:
                        print(f'{name}:\"{index}\"行没有{label}标签')
                else:
                    print(f'在{name}中没有找到{name}:\"{index}\"行')
        print(f'去除标签\"{label}\"完成')
    # 默认输出ProgramData，可用来输出搜索结果
    def Display(self,target:dict=False,namelist:list=False):
        if not target:
            target=self.ProgramData.copy()
        if not namelist:
            namelist=target.keys()
        for key in namelist:
            if key in target.keys():
                print(f'{key}:')
                # 完整输出DataFrame
                with pd.option_context('display.max_rows', None,    # 行数
                                       'display.max_columns', None, # 列数
                                       'display.width', None,       # 单元格长度
                                       'display.max_colwidth', None,# 不折叠单元格
                                        # 对齐表格(对不齐）
                                       # 'display.unicode.ambiguous_as_wide', True,
                                       # 'display.unicode.east_asian_width', True,
                                       # 'display.width',200
                                        ):
                    print(target[key].reset_index(drop=True,inplace=False))
            else:
                print(f'{key}不再目标范围内')
    # 按条件搜索，col是按搜索目标，0原文，1译文，2地址，3标签，4code。搜索条件为按*分割,target和返回值格式同ProgramData,notin为True，则搜索不含搜索目标的
    # BigSmall为true则不区分大小写
    def search(self,string:str,col:int,target:dict=False,namelist:list=False,notin:bool=False,BigSmall=False) ->dict:
        if col == 0:col = '原文'
        elif col == 1:col = '译文'
        elif col == 2:col = '地址'
        elif col == 3: col = '标签'
        elif col==4:col='code'
        string=string.split('*')
        res={}
        if not target:
            target=self.ProgramData.copy()
        if not namelist:
            namelist=target.keys()
        for name in namelist:
            if name in target.keys():
                DataFrame = target[name]
                if BigSmall:
                    temp=DataFrame.apply(lambda x: x.astype(str).str.lower())
                else:
                    temp=DataFrame.copy()
                for chara in string:
                    if BigSmall:
                        chara=chara.lower()
                    temp=temp[temp[col].str.contains(chara)]
                # 根据是否反选，返回未被改变大小写的dataframe
                if notin:
                    temp=DataFrame[~DataFrame.index.isin(temp.index)].dropna()
                else:
                    temp = DataFrame[DataFrame.index.isin(temp.index)].dropna()
                if len(list(temp.index)):
                    res.update({name:temp})
        return res
    # 搜索含A但不含B的，默认colB=colA
    def DoubleSearch(self,A:str,B:str,colA:int,colB:int=False,target:dict=False,namelist:list=False,BigSmall=False):
        res=self.search(A,colA,target,namelist,notin=False,BigSmall=BigSmall)
        if not colB:
            colB=colA
        return self.search(B,colB,res,namelist,True,BigSmall)
    # 替换,只能替换译文列
    def Replace(self,before:str,after:str,target:dict=False,namelist:list=False):
        if not target:
            target=self.ProgramData
        if not namelist:
            namelist=target.keys()
        for name in namelist:
            # DataFrame的replace不知道为什么就是换不掉
            for index in target[name].index:
                target[name].loc[index,'译文']=target[name].loc[index,'译文'].replace(before,after)
        return target
    # 根据搜索结果增减标签,add为True，添加标签,返回搜索结果
    def LabelBySearch(self,string:str,col:int,label:str,target:dict=False,namelist:list=False,notin:bool=False,BigSmall=False,add=True):
        res=self.search(string,col,target=target,namelist=namelist,notin=notin,BigSmall=BigSmall)
        target={}
        for name in res.keys():
            target.update({name:list(res[name].index)})
        if add:
            self.addlabel(target,label)
        else:
            self.removelabel(target,label)
        return res
    # 输出搜索结果,返回搜索结果
    def DisplayBySearch(self,string:str,col:int,target:dict=False,namelist:list=False,notin:bool=False,BigSmall=False):
        res=self.search(string,col,target=target,namelist=namelist,notin=notin,BigSmall=BigSmall)
        self.Display(res)
        return res
    # 将搜索结果导出到当前目录的单个xlsx中,返回搜索结果
    def OutputBySearch(self,string:str,col:int,target:dict=False,namelist:list=False,notin:bool=False,BigSmall=False,OutputName:str='SearchRes.xlsx'):
        res=self.search(string,col,target=target,namelist=namelist,notin=notin,BigSmall=BigSmall)
        if len (res):
            output=pd.concat(list(res.values()),axis=0)
            output.to_excel(OutputName, index=False)
            print(f'已将搜索结果保存为{OutputName}')
        else:
            print('搜索结果为空')
        return res
    # 将搜索结果导出到当前目录的单个json文件中,返回搜索结果
    def JsonBySearch(self,string:str,col:int,target:dict=False,namelist:list=False,notin:bool=False,BigSmall=False,OutputName:str='SearchRes.json'):
        res=self.search(string,col,target=target,namelist=namelist,notin=notin,BigSmall=BigSmall)
        if len(res):
            res = pd.concat(list(res.values()), axis=0)
            output=dict(zip(res['原文'],res['原文']))
            out = json.dumps(output, indent=4, ensure_ascii=False)
            with open(OutputName,'w',encoding='utf8') as f:
                print(out, file=f)
            print(f'已将结果导出为{OutputName}')
        else:
            print('搜索结果为空')
        return res
#######################################################预处理和后处理######################################################
    # 标签黑名单地址,标签为'BlackDir'。同时对其应用原文
    def LabelBlackDir(self):
        for i in self.BlackDir:
            res=self.LabelBySearch(i,2,'BlackDir')
            self.ApplyUntrs(res)
        for i in self.BlackCode:
            res = self.LabelBySearch(i, 4, 'BlackDir')
            self.ApplyUntrs(res)
        print('全部黑名单标记完成')
    # 标签名称为'Name',withoutx形如['Actors.json','Items.json','Skills.json'],可除外这些文件中的name，一般对应文件含对应对象的名字
    def LabelName(self,without:list=False):
        # target=self.search('BlackDir',3,notin=True) # 目标为不含'BlackDir'标签的行
        target=self.ProgramData
        if without:
            # 从target中依次除外without
            for i in without:
                target=self.search(i,2,target=target,notin=True)
        # 对剩下的标签'Name',不区分大小写搜索
        self.LabelBySearch('name',2,'Name',target=target,BigSmall=True)
    # 对名字标签并导出json文件
    def GetName(self,without:list=False):
        self.LabelName(without)
        self.JsonBySearch('Name',3,OutputName='Name.json')
    # 对target翻译应用原文
    def ApplyUntrs(self,target):
        for name in target.keys():
            if name in self.ProgramData.keys():
                for index in target[name].index:
                    if index in self.ProgramData[name].index:
                        self.ProgramData[name].loc[index,'译文']=index
        print('应用原文完成')
    # 对搜索结果应用原文
    def ApplyUntrs_BySearch(self,string:str,col:int,target:dict=False,namelist:list=False,notin:bool=False,BigSmall=False):
        res = self.search(string, col, target=target, namelist=namelist, notin=notin, BigSmall=BigSmall)
        self.ApplyUntrs(res)
    # 自动对游戏标题添加水印（地址为'System.json\gameTitle'的译文末尾添加mark）
    def AddMark(self,mark:str):
        if 'System.json' in self.ProgramData.keys():
            try:
                data=self.ProgramData['System.json']
                index=list(data[data['地址']==r'System.json\gameTitle'].index)[0]
                self.ProgramData['System.json'].loc[index,'译文']+=mark
                print('########################已添加水印########################')
            except Exception as e:
                print(traceback.format_exc())
                print(e)
                input(f'没有找到游戏标题，添加水印失败')
    # dnb用，分割原文译文的函数
    def __splitbychar(self,q, l):
        b = []
        if l[0] in q and l[1] in q:
            a = q.split(l[0])
            b.append(a[0])
            for i in range(1, len(a)):
                a[i] = l[0] + a[i]
                c = a[i].split(l[1])
                for j in range(0, len(c) - 1): b.append(c[j] + l[1])
                b.append(c[-1])
        return b
    # dnb用，处理原文包含文件名，但不等于文件名时，返回处理后译文
    def __dealin(self,untrs, trsed, filename):
        check = False
        for l in self.codewithnames:
            untrs_list = self.__splitbychar(untrs, l)
            trsed_list = self.__splitbychar(trsed, l)
            length = len(trsed_list)
            # 如果按某分隔符拆分结果长度不相等，不处理并单独导出
            if len(untrs_list) != length:
                self.need2check.update({untrs: trsed})
                return trsed
            elif length:
                # 如果拆分结果长度大于3，说明有不止一个分隔符，需导出确认
                if length > 3: check = True
                # 分隔符不相等
                if l[0] != l[1]:
                    # 被分隔符包裹的字符串只会出现在奇数位
                    for i in range(0, int(length / 2)):
                        # 如果文件名这一段在原文中，把对应位置的译文替换回原文
                        if filename in untrs_list[2 * i + 1]:
                            trsed_list[2 * i + 1] = untrs_list[2 * i + 1]
                else:
                    # 分隔符包裹的字符只会出现在4*i+2位
                    for i in range(0, int(length / 4)):
                        if filename in untrs_list[4 * i + 2]:
                            trsed_list[4 * i + 2] = untrs_list[4 * i + 2]
                # 将处理后的文本拼接好，等待下一循环
                trsed = ''
                for i in range(0, length): trsed += trsed_list[i]
        if check:
            self.need2check.update({untrs: trsed})
        return trsed
    # 处理文件名被翻译的问题
    def dnb(self,GameDir):
        print('开始修正文件名')
        temp=self.__ReadFolder(GameDir)
        files=[]
        self.need2check={}
        # 得到含有中日字符的文件名和不带后缀的形式
        for filename in temp:
            filename = filename.split('\\')[-1]
            if re.search(r'[\u4e00-\u9fa5\u3040-\u309f\u30a0-\u30ff\u4e00-\u9fa5]',filename):
                files.append(filename)
                files+=filename.split('.')[:-1]
        # 遍历
        for name in self.ProgramData.keys():
            for index in self.ProgramData[name].index:
                for filename in files:
                    # 原文等于文件名，译文替换回原文
                    if index== filename:
                        self.ProgramData[name].loc[index,'译文']=index
                    # 原文包含文件名时，在文件名被特定符号包裹的情况下，将被特定符号包裹的文本，按顺序替换回原文
                    elif filename in index:
                        self.ProgramData[name].loc[index,'译文']=self.__dealin(index,self.ProgramData[name].loc[index,'译文'],
                                                                             filename)
        print('########################修正文件名完成########################')
        # 如有需确认的文本，导出json文件
        if len(self.need2check):
            out = json.dumps(self.need2check, indent=4, ensure_ascii=False)
            with open('need2check.json', 'w', encoding='utf8') as f1:
                print(out, file=f1)
            print('已将可能需要人工修正的文本行导出到need2check.json')
    # 自动换行
    def AutoLineFeed(self,linelength:int):
        for name in self.ProgramData.keys():
            DataFrame=self.ProgramData[name]
            for index in DataFrame.index:
                data=DataFrame.loc[index,'译文']
                # 按换行符拆分译文
                lines = data.split('\n')
                res = ''
                for line in lines:
                    q = line
                    # 如果长度大于设定单行最大值，则进行拆分
                    while len(q) > linelength:
                        # 从字符串的第linelength-1个字符开始匹配中文汉字和英文字母，并返回其位置
                        n = re.search(r'[0-9a-zA-Z\u4e00-\u9fa5]', q[linelength:])
                        if n != None:
                            n = n.span()[0] + linelength
                        else:
                            break
                        res += q[:n] + '\n'
                        q = q[n:]
                    res += q + '\n'
                DataFrame.loc[index,'译文']= res.rstrip('\n')
            self.ProgramData[name]=DataFrame
    # 核对原文译文中，文本出现次数，若不同，单独导出。只导出原文至少出现一次的。需要一个名为checkdict.json的检查字典，格式为
    #{"要检查的原文":"对应的译文"}
    def checknum(self):
        count = 0
        # 加载检查字典，从长倒短排序
        try:
            with open('checkdict.json', 'r', encoding='utf8') as f:
                tempdict = json.load(f)
            # 按从长到短排序
            sortedkey = sorted(tempdict.keys(), key=lambda x: len(x), reverse=True)
            checkdict = {}
            for i in sortedkey:
                checkdict[i] = tempdict[i]
        except Exception as e:
            print(e)
            input('没有找到格式正确的checkdict文件，请确认路径设置是否正确以及文件是否符合json格式')
        res = {}
        fixdict = {}
        print('处理中，请稍候,根据checkdict的长度，可能会花费较长时间')
        for CheckUntrs in checkdict.keys():
            tempdict={}
            CheckTrsed=checkdict[CheckUntrs]
            for name in self.ProgramData.keys():
                DataFrame=self.ProgramData[name]
                for untrs in DataFrame.index:
                    trsed=DataFrame.loc[untrs,'译文']
                    if untrs.count(CheckUntrs)!=trsed.count(CheckTrsed) and untrs.count(CheckUntrs)>0:
                        tempdict.update({untrs:trsed})
            if len(tempdict):
                res.update({CheckUntrs: tempdict})
                fixdict.update({CheckUntrs: [CheckTrsed]})
                count += len(tempdict)
        out = json.dumps(res, indent=4, ensure_ascii=False)
        with open('checkres.json', 'w', encoding='utf8') as f1:
            print(out, file=f1)
        print('已将检查结果保存到checkres.json')
        # 询问是否覆盖fixdict.json
        overw = True
        if os.path.exists('fixdict.json'):
            inp = 0
            overw = False
            while inp not in ['y', 'n']:
                inp = input('fixdict.json已存在，请问要覆盖吗（y,n)\n').lower()
            if inp == 'y': overw = True
        if overw:
            self.tojson(fixdict, 'fixdict.json')
            print('对应的修正词典正确翻译导出为fixdict.json\n'
                  '修正词典的格式为：\n'
                  '{\n'
                  '"原文1":["正确译文","错误译文1","错误译文2"],\n'
                  '"原文2":["正确译文","错误译文1","错误译文2"]\n'
                  '}')
        print("########################核对完毕########################")
    # 根据核对结果，对翻译结果进行校正,只对核对结果做替换
    def fixnum(self):
        try:
            with open(r'checkres.json', 'r', encoding='utf8') as f:
                checkres = json.load(f)
        except Exception as e:
            print(e)
            input('没有找到格式正确的checkdict.json文件，请确认文件是否存在且符合json格式')
        try:
            with open(r'fixdict.json', 'r', encoding='utf8') as f:
                tempdict = json.load(f)
            sortedkey = sorted(tempdict.keys(), key=lambda x: len(x), reverse=True)
            fixdict = {}
            for i in sortedkey:
                fixdict[i] = tempdict[i]
        except Exception as e:
            print(e)
            input('没有找到格式正确的fixdict文件，请确认路径设置是否正确以及文件是否符合json格式')
        print('处理中，请稍候')
        TrsData={}  # 保存处理完毕的键值对
        for fixkey in list(fixdict.keys()):
            if fixkey in checkres.keys():
                righttrs = fixdict[fixkey][0]
                for untrs in checkres[fixkey].keys():
                    if fixkey in untrs:
                        for i in fixdict[fixkey][1:]:
                            # 如果相应文本之前已经替换过，则对TrsData内的操作
                            if untrs in TrsData.keys():
                                TrsData[untrs]=TrsData[untrs].replace(i, righttrs)
                            else:
                                checkres[fixkey][untrs] = checkres[fixkey][untrs].replace(i, righttrs)
                    if untrs not in TrsData.keys():
                        TrsData.update({untrs:checkres[fixkey][untrs]})
        self.InputFromJson(trsdata=TrsData)
        print('已应用修正结果')
    # 处理note中<SG説明:这类不能被翻译的文本，将其还原。特征是地址为含note，且包含数量相等的<和:
    # 这类文本被翻译，通常会导致图鉴等不显示
    def DNoteB(self):
        print('正在处理可能存在的note问题')
        res=self.search('note',2)
        for name in res.keys():
            DataFrame=res[name]
            for untrs in DataFrame.index:
                trsed=DataFrame.loc[untrs,'译文']
                l=['<',':'] # 分隔符
                if untrs.count(l[0])==untrs.count(l[1]) and untrs.count(l[0])==trsed.count(l[0]) and\
                    untrs.count(l[1])==trsed.count(l[1]):
                    # 将文本按照分隔符拆分
                    untrs_list = self.__splitbychar(untrs, l)
                    trsed_list = self.__splitbychar(trsed, l)
                    length = len(trsed_list)
                    # 如果按某分隔符拆分结果长度不相等，不处理
                    if len(untrs_list) != length:
                        continue
                    elif length:
                        # 被分隔符包裹的字符串只会出现在奇数位
                        for i in range(0, int(length / 2)):
                            trsed_list[2 * i + 1] = untrs_list[2 * i + 1]
                        # 将处理后的文本拼接好，等待下一循环
                        trsed = ''
                        for i in range(0, length): trsed += trsed_list[i]
                DataFrame.loc[untrs,'译文']=trsed
            # 将处理后文本导入到工程数据
            self.InputFromDataFrame(DataFrame,[name])
        print('########################note处理完毕########################')
#########################################################一键处理#########################################################
    # 读取游戏并打好预设标签，获取人名，保存数据，参数为游戏目录和保存路径，保存路径需是已存在的文件夹
    def FromGame(self,GameDir,path):
        self.ReadGame(GameDir)
        self.GetName(self.NameWithout)
        self.Save(path)
    # 注入游戏，自动处理文件名问题，如有水印(如有），打水印，参数为游戏根目录,翻译数据路径和注入翻译后的json文件保存目录，mark为水印
    def ToGmae(self,GameDir,path,OutputPath,mark:str=False):
        self.InputFromeXlsx(path)
        self.dnb(GameDir)
        self.DNoteB()
        if mark:
            self.AddMark(mark)
        self.InjectGame(GameDir,OutputPath)
# 从路径读取config.json，并初始化翻译器，随后读取翻译数据,返回翻译器
def Jr_Tpp_LOAD(path):
    try:
        with open(path+'\\'+'config.json','r',encoding='utf8') as f:
            config=json.load(f)
    except Exception as e:
        print(traceback.format_exc())
        print(e)
        print('请确保读取路径内包含格式正确的config.json文件')
    jrtpp=Jr_Tpp(config,path)
    return jrtpp
def readconfig():
    try:
        yaml = YAML(typ='safe')
        with open('config.yaml', 'r', encoding='utf8') as f:
            config = yaml.load(f)
        # 确保config中数据齐全
        game_path=config['game_path']
        save_path=config['save_path']
        translation_path=config['translation_path']
        mark=config['mark']
        NameWithout=config['NameWithout']
        ReadCode=config['ReadCode']
        BlackDir=config['BlackDir']
        BlackCode=config['BlackCode']
        BlackFiles=config['BlackFiles']
        codewithnames=config['codewithnames']
        output_path=config['output_path']
        line_length=config['line_length']
    except Exception as e:
        print(e)
        input('没有找到格式正确的config.yaml文件，请确保其存在于与exe同级文件夹内')
    return config
if __name__ == '__main__':
    config=readconfig()
    startpage='1.一键读取游戏数据并保存\n' \
              '2.加载翻译工程\n'
    key=['1','2']
    try:
        while 1:
            res=0
            while res not in key:
                res = input(startpage)
            if res=='1':
                pj=Jr_Tpp(config)
                pj.FromGame(config['game_path'],config['save_path'])
                input('已成功读取游戏数据，提取到的名字保存在Name.json中\n'
                      '请在翻译完名字以后，将其导入到ainiee的提示词典中\n'
                      '然后翻译{}\\data中的xlsx文件\n'.format(config['save_path']))
            else:
                pj = Jr_Tpp_LOAD(config['save_path'])
            mainpage = '1.一键注入翻译\n' \
                       '2.自动换行（换行后不会自动保存，也不会自动注入，不推荐在没有没备份的情况下保存)\n' \
                       '3.保存翻译工程\n' \
                       '4.加载翻译工程\n' \
                       '5.重新加载配置文件\n'
            while 1:
                res=0
                while res not in ['1','2','3','4','5']:
                    res = input(mainpage)
                if res=='1':
                    pj.ToGmae(config['game_path'],config['translation_path'],config['output_path'],config['mark'])
                elif res=='2':
                    pj.AutoLineFeed(config['line_length'])
                elif res=='3':
                    pj.Save(config['save_path'])
                elif res=='4':
                    pj = Jr_Tpp_LOAD(config['save_path'])
                elif res=='5':
                    config=readconfig()
                    print('已重新加载配置文件')
    except Exception as e:
        print(traceback.format_exc())
        print(e)
        input('发生错误，请上报bug')
    # 读
    # test=Jr_Tpp(config)
    # test.FromGame(config['GameDir'],'data')
    # test.OutputBySearch('vents*name',2)


    #
    # test.ReadGame(config['GameDir'])
    # test.GetName()
    # test.InputFromJson(path=r'res/TrsData.json')
    # test.Save('data')
    # 写
    # test=Jr_Tpp_LOAD('data')
    # translation_path=r'D:\ggsddu\old\QFT\system\mytrs\jt++\jt++\ainiee'
    # outputpath='D:\ggsddu\old\QFT\system\mytrs\jt++\output'
    # test.ToGmae(config['GameDir'],translation_path,outputpath,config['mark'])
    # test.Save('data')
    # test.DNoteB()
    # test.Save('data')
    # # test.DisplayBySearch('name',2,BigSmall=True)
    # print(test.GetFileNames())
    # test.Display(namelist=['Actors.json'])