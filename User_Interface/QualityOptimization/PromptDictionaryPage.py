
import os
import json

import openpyxl
from rich import print
from PyQt5.Qt import Qt
from PyQt5.QtWidgets import QFrame
from PyQt5.QtWidgets import QWidget
from PyQt5.QtWidgets import QFileDialog
from PyQt5.QtWidgets import QHeaderView
from PyQt5.QtWidgets import QVBoxLayout
from PyQt5.QtWidgets import QTableWidgetItem

from qfluentwidgets import Action
from qfluentwidgets import InfoBar
from qfluentwidgets import InfoBarPosition
from qfluentwidgets import CommandBar
from qfluentwidgets import FluentIcon
from qfluentwidgets import MessageBox
from qfluentwidgets import TableWidget

from Widget.SpinCard import SpinCard
from Widget.ComboBoxCard import ComboBoxCard
from Widget.CommandBarCard import CommandBarCard
from Widget.SwitchButtonCard import SwitchButtonCard

class PromptDictionaryPage(QFrame):
    
    DEFAULT = {
        "prompt_dict_switch": True,
        "User_Dictionary2": {
            "ダリヤ": {
                "translation": "达莉雅",
                "info": "女性的名字"
            },
        },
    }

    def __init__(self, text: str, parent = None, configurator = None):
        super().__init__(parent = parent)

        self.setObjectName(text.replace(" ", "-"))
        self.configurator = configurator

        # 载入配置文件
        config = self.load_config()
        config = self.save_config(config)

        # 设置主容器
        self.container = QVBoxLayout(self)
        self.container.setSpacing(8)
        self.container.setContentsMargins(24, 24, 24, 24) # 左、上、右、下

        # 添加控件
        self.add_widget_header(self.container, config)
        self.add_widget_body(self.container, config)
        self.add_widget_footer(self.container, config, parent)

    # 载入配置文件
    def load_config(self) -> dict:
        config = {}

        if os.path.exists(os.path.join(self.configurator.resource_dir, "config.json")):
            with open(os.path.join(self.configurator.resource_dir, "config.json"), "r", encoding = "utf-8") as reader:
                config = json.load(reader)
        
        return config

    # 保存配置文件
    def save_config(self, new: dict) -> None:
        path = os.path.join(self.configurator.resource_dir, "config.json")
        
        # 读取配置文件
        if os.path.exists(path):
            with open(path, "r", encoding = "utf-8") as reader:
                old = json.load(reader)
        else:
            old = {}

        # 修改配置文件中的条目：如果条目存在，这更新值，如果不存在，则设置默认值
        for k, v in self.DEFAULT.items():
            if not k in new.keys():
                old[k] = v
            else:
                old[k] = new[k]

        # 写入配置文件
        with open(path, "w", encoding = "utf-8") as writer:
            writer.write(json.dumps(old, indent = 4, ensure_ascii = False))

        return old

    # 头部
    def add_widget_header(self, parent, config):
        def widget_init(widget):
            widget.setChecked(config.get("prompt_dict_switch"))
            
        def widget_callback(widget, checked: bool):
            config["prompt_dict_switch"] = checked
            self.save_config(config)

        parent.addWidget(
            SwitchButtonCard(
                "指令词典", 
                "通过构建词典指令来引导模型翻译，可实现统一翻译、矫正人称属性等功能 (不支持 Sakura v0.9 模型)",
                widget_init,
                widget_callback,
            )
        )

    # 主体
    def add_widget_body(self, parent, config):
        self.table = TableWidget(self)
        parent.addWidget(self.table)

        # 启用边框并设置圆角
        self.table.setBorderRadius(4)
        self.table.setBorderVisible(True)

        self.table.setWordWrap(False)
        self.table.setRowCount(12)
        self.table.setColumnCount(3)
        self.table.resizeRowsToContents() # 设置行高度自适应内容
        self.table.resizeColumnsToContents() # 设置列宽度自适应内容
        self.table.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch) # 撑满宽度

        # 设置水平表头并隐藏垂直表头
        self.table.verticalHeader().hide()
        self.table.setHorizontalHeaderLabels(
            [
                "原文",
                "译文",
                "描述",
            ],
        )

        # 向表格更新数据
        self.update_to_table(self.table, config)

    # 底部
    def add_widget_footer(self, parent, config, window):
        self.command_bar_card = CommandBarCard()
        parent.addWidget(self.command_bar_card)
        
        # 添加命令
        self.add_command_bar_action_01(self.command_bar_card)
        self.add_command_bar_action_02(self.command_bar_card)
        self.command_bar_card.addSeparator()
        self.add_command_bar_action_03(self.command_bar_card)
        self.add_command_bar_action_04(self.command_bar_card)
        self.command_bar_card.addSeparator()
        self.add_command_bar_action_05(self.command_bar_card)
        self.add_command_bar_action_06(self.command_bar_card, window)

    # 向表格更新数据
    def update_to_table(self, table, config):
        datas = []
        user_dictionary = config.get("User_Dictionary2", {})
        table.setRowCount(max(12, len(user_dictionary)))
        for k, v in user_dictionary.items():
            datas.append(
                [k.strip(), v.get("translation", "").strip(), v.get("info", "").strip()]
            )
        for row, data in enumerate(datas):
            for col, v in enumerate(data):
                item = QTableWidgetItem(v)
                item.setTextAlignment(Qt.AlignCenter)
                table.setItem(row, col, item)

    # 从表格更新数据
    def update_from_table(self, table, config):
        config["User_Dictionary2"] = {}
        
        for row in range(table.rowCount()):
            data_str = table.item(row, 0)
            data_dst = table.item(row, 1)
            data_info = table.item(row, 2)

            # 判断是否有数据
            if data_str == None or data_dst == None:
                continue
            
            data_str = data_str.text().strip()
            data_dst = data_dst.text().strip()
            data_info = data_info.text().strip() if data_info != None else ""

            # 判断是否有数据
            if data_str == "" or data_dst == "":
                continue

            config["User_Dictionary2"][data_str] = {
                "translation": data_dst,
                "info": data_info,
            }

        return config

    # 导入
    def add_command_bar_action_01(self, parent):

        def load_json_file(path):
            dictionary = {}
            
            inputs = []
            with open(path, "r", encoding = "utf-8") as reader:
                inputs = json.load(reader)

            if isinstance(inputs, list) and len(inputs) > 0:
                for v in inputs:
                    # 标准术语表
                    # [
                    #     {
                    #         "srt": "ダリヤ",
                    #         "dst": "达莉雅",
                    #         "info": "女性的名字"
                    #     }
                    # ]
                    if isinstance(v, dict) and v.get("srt", "") != "" and v.get("dst", "") != "":
                        dictionary[v.get("srt", "").strip()] = {
                            "translation": v.get("dst", "").strip(),
                            "info": v.get("info", "").strip(),
                        }
                    
                    # Paratranz的术语表
                    # [
                    #   {
                    #     "id": 359894,
                    #     "createdAt": "2024-04-06T18:43:56.075Z",
                    #     "updatedAt": "2024-04-06T18:43:56.075Z",
                    #     "updatedBy": null,
                    #     "pos": "noun",
                    #     "uid": 49900,
                    #     "term": "アイテム",
                    #     "translation": "道具",
                    #     "note": "",
                    #     "project": 9841,
                    #     "variants": []
                    #   }
                    # ]
                    if isinstance(v, dict) and v.get("term", "") != "" and v.get("translation", "") != "":
                        dictionary[v.get("term", "").strip()] = {
                            "translation": v.get("translation", "").strip(),
                            "info": "",
                        }
            elif isinstance(inputs, dict):
                # 普通 KV 格式
                # [
                #     "ダリヤ": "达莉雅"
                # ]
                for k, v in inputs.items():
                    if isinstance(v, str) and k != "" and v != "":
                        dictionary[k.strip()] = {
                            "translation": v.strip(),
                            "info": "",
                        }

            return dictionary
            
        def load_xlsx_file(path):
            dictionary = {}

            sheet = openpyxl.load_workbook(path).active
            for row in range(2, sheet.max_row + 1): # 第一行是标识头，第二行才开始读取
                cell_value1 = sheet.cell(row=row, column=1).value # 第N行第一列的值
                cell_value2 = sheet.cell(row=row, column=2).value # 第N行第二列的值
                cell_value3 = sheet.cell(row=row, column=3).value # 第N行第三列的值

                if cell_value1 != "" and cell_value2 != "":
                    dictionary[cell_value1.strip()] = {
                        "translation": cell_value2.strip(),
                        "info": cell_value3.strip(),
                    }

            return dictionary
        
        def callback():
            # 选择文件
            path, _ = QFileDialog.getOpenFileName(None, "选择文件", "", "json files (*.json);;xlsx files (*.xlsx)")
            if path == None or path == "":
                return

            # 获取文件后缀
            file_suffix = path.split(".")[-1].lower()

            datas = []
            if file_suffix == "json":
                datas = load_json_file(path)
                
            if file_suffix == "xlsx":
                datas = load_xlsx_file(path)

            # 读取配置文件
            config = self.load_config()
            config["User_Dictionary2"].update(datas)

            # 保存配置文件
            config = self.save_config(config)

            # 向表格更新数据
            config = self.update_to_table(self.table, config)

            # 弹出提示
            InfoBar.success(
                title = "",
                content = "文件数据已导入 ...",
                parent = self,
                duration = 2000,
                orient = Qt.Horizontal,
                position = InfoBarPosition.TOP,
                isClosable = True,
            )

        parent.addAction(
            Action(FluentIcon.DOWNLOAD, "导入", parent, triggered = callback),
        )
        
    # 导出
    def add_command_bar_action_02(self, parent):
        def callback():
            # 读取配置文件
            config = self.load_config()

            # 从表格更新数据
            config = self.update_from_table(self.table, config)

            # 整理数据
            datas = []
            user_dictionary = config.get("User_Dictionary2", {})
            for k, v in user_dictionary.items():
                datas.append(
                    {
                        "srt": k,
                        "dst": v.get("translation", ""),
                        "info": v.get("info", ""),
                    }
                )

            # 选择文件导出路径
            path = QFileDialog.getExistingDirectory(None, "Select Directory", "")
            if path == None or path == "":
                return

            # 导出文件
            with open(os.path.join(path, "导出_指令词典.json"), "w", encoding = "utf-8") as writer:
                writer.write(json.dumps(datas, indent = 4, ensure_ascii = False))

            # 弹出提示
            InfoBar.success(
                title = "",
                content = f"表格数据已导出为 导出_指令词典.json ...",
                parent = self,
                duration = 2000,
                orient = Qt.Horizontal,
                position = InfoBarPosition.TOP,
                isClosable = True,
            )

        parent.addAction(
            Action(FluentIcon.SHARE, "导出", parent, triggered = callback),
        )
        
    # 添加新行
    def add_command_bar_action_03(self, parent):
        def callback():
            # 添加新行
            self.table.setRowCount(self.table.rowCount() + 1)

            # 弹出提示
            InfoBar.success(
                title = "",
                content = "新行已添加 ...",
                parent = self,
                duration = 2000,
                orient = Qt.Horizontal,
                position = InfoBarPosition.TOP,
                isClosable = True,
            )

        parent.addAction(
            Action(FluentIcon.ADD_TO, "添加新行", parent, triggered = callback),
        )

    # 移除空行
    def add_command_bar_action_04(self, parent):
        def callback():
            # 从表格更新数据，生成一个临时的配置文件
            config = self.update_from_table(self.table, {})

            # 清空表格
            self.table.clearContents()

            # 向表格更新数据
            self.update_to_table(self.table, config)

            # 弹出提示
            InfoBar.success(
                title = "",
                content = "空行已移除 ...",
                parent = self,
                duration = 2000,
                orient = Qt.Horizontal,
                position = InfoBarPosition.TOP,
                isClosable = True,
            )

        parent.addAction(
            Action(FluentIcon.BROOM, "移除空行", parent, triggered = callback),
        )

    # 保存
    def add_command_bar_action_05(self, parent):
        def callback():
            # 读取配置文件
            config = self.load_config()

            # 从表格更新数据
            config = self.update_from_table(self.table, config)

            # 保存配置文件
            config = self.save_config(config)

            # 弹出提示
            InfoBar.success(
                title = "",
                content = "表格数据已保存 ...",
                parent = self,
                duration = 2000,
                orient = Qt.Horizontal,
                position = InfoBarPosition.TOP,
                isClosable = True,
            )

        parent.addAction(
            Action(FluentIcon.SAVE, "保存", parent, triggered = callback),
        )
        
    # 重置
    def add_command_bar_action_06(self, parent, window):
        def callback():
            message_box = MessageBox("警告", "是否确认重置为默认数据 ... ？", window)
            message_box.yesButton.setText("确认")
            message_box.cancelButton.setText("取消")

            if not message_box.exec():
                return

            # 清空表格
            self.table.clearContents()

            # 读取配置文件
            config = self.load_config()

            # 加载默认设置
            for k, v in self.DEFAULT.items():
                config[k] = v

            # 保存配置文件
            config = self.save_config(config)

            # 向表格更新数据
            self.update_to_table(self.table, config)

            # 弹出提示
            InfoBar.success(
                title = "",
                content = "表格数据已重置 ...",
                parent = self,
                duration = 2000,
                orient = Qt.Horizontal,
                position = InfoBarPosition.TOP,
                isClosable = True,
            )

        parent.addAction(
            Action(FluentIcon.DELETE, "重置", parent, triggered = callback),
        )